from openpyxl import load_workbook


def read_excel_file(filename: str, sheet_names: tuple = None) -> {tuple}:
    workbook = load_workbook(filename=filename, read_only=True, data_only=True)
    if not sheet_names:
        sheet_names = workbook.sheetnames
    all_data = {}
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        all_data[sheet_name] = data
    workbook.close()
    return all_data


