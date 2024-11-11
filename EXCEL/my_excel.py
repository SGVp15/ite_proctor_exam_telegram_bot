import re

from openpyxl import load_workbook


def read_excel_file(filename: str, sheet_names: tuple=None) -> {tuple}:
    workbook = load_workbook(filename=filename, read_only=True, data_only=True)
    if sheet_names is None:
        sheet_names = workbook.sheetnames
    all_data = {}
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        all_data[sheet_name] = data
    workbook.close()
    return all_data


def clean_str(s):
    if type(s) is str:
        s = s.strip()
        s = s.replace(',', ', ')
        s = re.sub(r'\s{2,}', ' ', s)
        if s in ('None', '#N/A'):
            s = ''
    return s
