import json
import os
import re

import openpyxl
from openpyxl.utils import get_column_letter

from Question import Question


def get_all_questions_from_excel_file(file: str) -> [Question]:
    wb = openpyxl.load_workbook(filename=f'{file}', data_only=True)
    page_name = wb.sheetnames
    page_name = str(page_name[0])

    map_excel = {}
    max_column = wb[page_name].max_column
    for col in range(1, max_column + 1):
        map_excel[read_excel(wb, page_name, get_column_letter(col), 1)] = get_column_letter(col)
    column_id_question = map_excel['Код вопроса']
    column_category_question = map_excel['Раздел курса']
    column_box_question = map_excel['Блок вопросов']
    column_enable_question = map_excel.get('Действующий 1-да, 0-нет', 'L')
    column_image = map_excel.get('Рисунок', 'C')

    column_a = 'a'
    column_main = 'b'

    all_questions = []
    num_q = 0

    i = 0
    max_row = wb[page_name].max_row
    while i <= max_row:
        i += 1
        a = read_excel(wb, page_name, column_a, i + 1)

        if a not in ('a', 'а', 'A', 'А'):
            continue

        b = read_excel(wb, page_name, column_a, i + 2).lower()
        c = read_excel(wb, page_name, column_a, i + 3).lower()
        d = read_excel(wb, page_name, column_a, i + 4).lower()

        if b not in ('b', 'в') or c not in ('c', 'с') or d != 'd':
            print(f'[Error] row : {i}, {file=}')
            continue

        is_enable_question = read_excel(wb, page_name, column_enable_question, i)
        if is_enable_question in ('1', 1):
            q = Question()
            q.text_question = read_excel(wb, page_name, column_main, i)
            if q.text_question in ('', None):
                continue
            q.id_question = read_excel(wb, page_name, column_id_question, i)
            q.box_question = read_excel(wb, page_name, column_box_question, i)

            image = read_excel(wb, page_name, column_image, i)
            if image is not None:
                image_pattern = r'\s*([A-Яа-я\w\d _\-]+\.\w+)\s*'
                if re.search(image_pattern, image):
                    q.image = re.findall(pattern=image_pattern, string=image)[0]
                    q.image = os.path.join(os.path.dirname(file), q.image.strip())
            else:
                q.image = ''

            q.ans_a = read_excel(wb, page_name, column_main, i + 1)
            q.ans_b = read_excel(wb, page_name, column_main, i + 2)
            q.ans_c = read_excel(wb, page_name, column_main, i + 3)
            q.ans_d = read_excel(wb, page_name, column_main, i + 4)
            if None in (q.ans_a, q.ans_b, q.ans_c, q.ans_d):
                continue
            num_q += 1
            q.category = read_excel(wb, page_name, column_category_question, i)

            all_questions.append(q)
        i += 4

    return all_questions


def save_json_questions(path_questions, all_questions):
    file_json = os.path.join(path_questions, f'{all_questions[0].exam}.json')
    with open(file_json, 'w', encoding='utf-8') as f:
        f.write(json.dumps(all_questions))


def read_excel(excel, page_name, column, row):
    sheet_ranges = excel[page_name]
    v = sheet_ranges[f'{column}{row}'].value
    if v is None:
        return None
    value = str(v).strip()
    if value == '':
        return None
    return value
